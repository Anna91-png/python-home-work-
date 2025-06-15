import csv


def read_transactions_from_csv(file_path):
    transactions = []
    with open(file_path, newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            transactions.append(dict(row))
    return transactions


if __name__ == "__main__":
    # Здесь тестируем функции
    transactions = read_transactions_from_csv("../data/transactions.csv")
    print(transactions)
    print(type(transactions))
    if transactions:
        print(type(transactions[0]))

import openpyxl
import os

def read_transactions_from_excel(file_path):
    transactions = []
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        transaction = dict(zip(headers, row))
        transactions.append(transaction)
    return transactions

if __name__ == "__main__":
    print("Текущая рабочая директория:", os.getcwd())
    file_path = "../data/transactions_excel (1).xlsx"
    if not os.path.isfile(file_path):
        print(f"Файл не найден: {file_path}")
    else:
        transactions = read_transactions_from_excel(file_path)
        print(transactions)