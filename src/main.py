import json
import csv
from typing import List, Dict
import sys

import re
from src.bank_search import process_bank_search
from src.bank_operation import process_bank_operations

try:
    import openpyxl
except ImportError:
    openpyxl = None

AVAILABLE_STATUSES = ["EXECUTED", "CANCELED", "PENDING"]
CURRENCIES = ["руб.", "RUB", "руб", "RUB.", "Rur", "RUR"]

def load_json_transactions(filename: str) -> List[Dict]:
    with open(filename, encoding="utf-8") as f:
        return json.load(f)

def load_csv_transactions(filename: str) -> List[Dict]:
    with open(filename, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return list(reader)

def load_xlsx_transactions(filename: str) -> List[Dict]:
    if not openpyxl:
        print("Для работы с XLSX файлами требуется пакет openpyxl.")
        sys.exit(1)
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    data = []
    headers = [str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(headers, row)))
    return data

def filter_by_status(data: List[Dict], status: str) -> List[Dict]:
    # Учитываем только точное совпадение по приведенному к верхнему регистру статусу
    return [op for op in data if str(op.get('status', '')).upper() == status.upper()]

def sort_by_date(data: List[Dict], reverse: bool = False) -> List[Dict]:
    def parse_date(op):
        # Попробуйте извлечь дату из поля 'date', 'datetime', 'Дата', иначе None
        for key in ['date', 'datetime', 'Дата']:
            if key in op:
                # Пример: '2019-08-26T10:50:58.294041'
                val = op[key]
                if isinstance(val, str):
                    try:
                        # Стандарт ISO, если есть T
                        if "T" in val:
                            return val.split("T")[0]
                        # Если формат DD.MM.YYYY
                        if "." in val:
                            parts = val.split()
                            for part in parts:
                                if re.match(r"\d{2}\.\d{2}\.\d{4}", part):
                                    return "-".join(reversed(part.split(".")))
                        # Если просто YYYY-MM-DD
                        if "-" in val:
                            return val
                    except Exception:
                        pass
        return ""
    # Фильтруем пустые даты в конец
    return sorted(data, key=lambda op: parse_date(op) or "9999-12-31", reverse=reverse)

def filter_by_rub(data: List[Dict]) -> List[Dict]:
    # Поиск по ключам 'currency', 'amountCurrency', 'sum', 'Сумма', 'operationAmount'
    def is_rub(op):
        # Поиск по currency
        for key in ['currency', 'amountCurrency', 'sum', 'Сумма', 'operationAmount']:
            val = op.get(key)
            if not val:
                continue
            valstr = str(val)
            for cur in CURRENCIES:
                if cur.lower() in valstr.lower():
                    return True
        # Иногда валюта может быть отдельным полем
        if "currency" in op and str(op["currency"]).lower() in ["rub", "руб.", "rur"]:
            return True
        # Иногда в описании суммы
        if "description" in op and any(cur.lower() in str(op["description"]).lower() for cur in CURRENCIES):
            return True
        return False
    return [op for op in data if is_rub(op)]

def pretty_print_transactions(data: List[Dict]):
    print(f"\nВсего банковских операций в выборке: {len(data)}\n")
    for op in data:
        # Дата
        date = ""
        for key in ['date', 'Дата', 'datetime']:
            if key in op:
                raw = op[key]
                # Преобразуем дату к виду DD.MM.YYYY если можем
                if isinstance(raw, str):
                    if raw.count("-") == 2 and "T" in raw:
                        date = ".".join(reversed(raw.split("T")[0].split("-")))
                    elif raw.count(".") == 2:
                        date = raw.split()[0]
                    elif raw.count("-") == 2:
                        date = ".".join(reversed(raw.split("-")))
                    else:
                        date = raw
                else:
                    date = str(raw)
                break
        # Описание
        description = op.get("description") or op.get("Описание") or ""
        print(f"{date} {description}")
        # Счет/карта/отправитель/получатель
        for key in ["from", "from_", "отправитель"]:
            if key in op:
                print(str(op[key]), end=" -> ")
        for key in ["to", "получатель"]:
            if key in op:
                print(str(op[key]))
        # Сумма и валюта
        amount, currency = "", ""
        for key in ["amount", "sum", "Сумма", "operationAmount"]:
            if key in op:
                val = str(op[key])
                amount = val
                break
        for key in ["currency", "amountCurrency", "валюта"]:
            if key in op:
                currency = str(op[key])
                break
        # Иногда сумма и валюта в одном поле
        if amount and not any(cur in amount for cur in CURRENCIES):
            print(f"Сумма: {amount} {currency}\n")
        else:
            print(f"Сумма: {amount}\n")

def main():
    print("Привет! Добро пожаловать в программу работы с банковскими транзакциями.")
    print("Выберите необходимый пункт меню:")
    print("1. Получить информацию о транзакциях из JSON-файла")
    print("2. Получить информацию о транзакциях из CSV-файла")
    print("3. Получить информацию о транзакциях из XLSX-файла")
    data = []
    while True:
        choice = input()
        if choice == "1":
            print("Для обработки выбран JSON-файл.")
            filename = input("Введите имя JSON-файла (например, data.json): ")
            try:
                data = load_json_transactions(filename)
                break
            except Exception:
                print("Ошибка чтения файла. Попробуйте снова.")
        elif choice == "2":
            print("Для обработки выбран CSV-файл.")
            filename = input("Введите имя CSV-файла (например, data.csv): ")
            try:
                data = load_csv_transactions(filename)
                break
            except Exception:
                print("Ошибка чтения файла. Попробуйте снова.")
        elif choice == "3":
            print("Для обработки выбран XLSX-файл.")
            filename = input("Введите имя XLSX-файла (например, data.xlsx): ")
            try:
                data = load_xlsx_transactions(filename)
                break
            except Exception:
                print("Ошибка чтения файла. Попробуйте снова.")
        else:
            print("Некорректный пункт меню. Попробуйте снова.")

    # Выбор статуса
    while True:
        print("Введите статус, по которому необходимо выполнить фильтрацию. ")
        print("Доступные для фильтровки статусы: EXECUTED, CANCELED, PENDING")
        status = input()
        status_up = status.upper()
        if status_up in AVAILABLE_STATUSES:
            data = filter_by_status(data, status_up)
            print(f'Операции отфильтрованы по статусу "{status_up}"')
            break
        else:
            print(f'Статус операции "{status}" недоступен.')

    # Если после фильтрации пусто
    if not data:
        print("Не найдено ни одной транзакции, подходящей под ваши условия фильтрации")
        return

    # Сортировка по дате
    print("Отсортировать операции по дате? Да/Нет")
    sort_q = input().strip().lower()
    if sort_q in ["да", "yes", "y"]:
        print("Отсортировать по возрастанию или по убыванию?")
        order = input().strip().lower()
        if "возрастан" in order or "asc" in order or "+" in order:
            data = sort_by_date(data, reverse=False)
        else:
            data = sort_by_date(data, reverse=True)

    # Фильтрация по рублям
    print("Выводить только рублевые транзакции? Да/Нет")
    rub_q = input().strip().lower()
    if rub_q in ["да", "yes", "y"]:
        data = filter_by_rub(data)
        if not data:
            print("Не найдено ни одной транзакции, подходящей под ваши условия фильтрации")
            return

    # Фильтрация по слову в описании
    print("Отфильтровать список транзакций по определенному слову в описании? Да/Нет")
    desc_q = input().strip().lower()
    if desc_q in ["да", "yes", "y"]:
        search = input("Введите слово для поиска в описании: ").strip()
        data = process_bank_search(data, search)
        if not data:
            print("Не найдено ни одной транзакции, подходящей под ваши условия фильтрации")
            return

    print("Распечатываю итоговый список транзакций...")
    if data:
        pretty_print_transactions(data)
    else:
        print("Не найдено ни одной транзакции, подходящей под ваши условия фильтрации")

if __name__ == "__main__":
    main()